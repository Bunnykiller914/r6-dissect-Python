import json
import pandas as pd
import os
import openpyxl
from openpyxl.styles import Border, Side, Font
import matplotlib.pyplot as plt


############################ Future potential work ############################
#have objs, not doing anything with though
#new sheet for each team, with their objs and win % with each
#check for dissect from .git, download if not there
#figure out updated rating eq from siege.gg
#use pulic replays from SI23 to calculate rating equation
#W/L and OBJ weird bc of "DefuserDisableComplete" and "DefuserPlantComplete"
#assists exist now
#1vx should work now

#read in large amount of matches, make a sheet for each team with the stats from every match played

################################################################################

os.chdir(os.path.dirname(os.path.abspath(__file__))) #change directory to the folder the program is in

#check if the folders exist, if not create them
checkfolders=False
if "MatchReplays" not in os.listdir():
    os.mkdir("MatchReplays")
    checkfolders=True
if "Jsons" not in os.listdir():
    os.mkdir("Jsons")
    checkfolders=True
if "Outputs" not in os.listdir():
    os.mkdir("Outputs")
    checkfolders=True
if "Dissected" not in os.listdir("Jsons"):
    os.mkdir("Jsons\\Dissected")
    checkfolders=True
if "Other" not in os.listdir("Jsons"):
    os.mkdir("Jsons\\Other")
    checkfolders=True
if "Dissected" not in os.listdir("MatchReplays"):
    os.mkdir("MatchReplays\\Dissected")
    checkfolders=True

if checkfolders==True:
    print("Folders created, please move match replays to the MatchReplays folder and run the program again")
    exit()

dissect=False
if input("Do you want to dissect a new match? (y/n) ").lower()=="y": #dissect a new match
    dissect=True
    os.chdir("MatchReplays")
    folders=next(os.walk('.'))[1]
    foldersfiltered=[]
    for item in folders:
        if item != ".vscode" and item != ".git" and item != "__pycache__" and item != "venv" and item != "Dissected" and item != "Ranked":
            foldersfiltered.append(item)
    for selection in foldersfiltered:
        print(foldersfiltered)
        choice=input("Do you want to dissect "+selection+"? (y/n) ").lower()
        if choice=="y":
            os.chdir(selection)
            break
        else:
            continue
    matches=next(os.walk('.'))[1]
    os.chdir('..')
    os.chdir('..')
    files=os.listdir("Jsons") #existing number of jsons
    skips=0
    for match in matches:
        #if match already exists, skip
        if match+".json" in files:
            print(""+match+" already exists in Jsons folder. Skipping...")
            skips+=1
            continue
        try:
            folderpath="MatchReplays\\"+selection+"\\"+match
        except:
            print("Error: Match not found, put a match in the MatchReplays folder and try again")
            print("Exiting...")
            exit()
        savepath="Jsons"+"\\"+match+".json"
        command="start cmd /c r6-dissect "+folderpath+" -o "+savepath+""
        os.system(command)  #creates json and closes
    while True:
        newfiles=os.listdir("Jsons")
        if len(newfiles)==len(files)+len(matches)-skips: #if the files are created, move on
            break
        else:
            continue
    while True:
        for item in os.listdir("Jsons"):
            if item != "Dissected" and item != "Ranked" and item != "Other" and item != "Project":
                while True:
                    file_stats = os.stat("Jsons\\"+item)
                    if file_stats.st_size < 1000:
                        continue
                    else:
                        break
            else:
                continue
        break
    os.rename("MatchReplays\\"+selection, "MatchReplays\\Dissected\\"+selection)
    print("Dissection complete")

path = os.getcwd()
path+='\\Jsons'
dir_list=os.listdir(path)
numberofmatches=0
matcheslist=[]
jsonlist=[]
othercount=0
for item in dir_list: #check if the file is a json
    if item.endswith('.json'):
        jsonlist.append(item)
    else:
        othercount+=1
notcomplist=[]
# for item in jsonlist: #check if the match is a custom game
#     with open(path+'\\'+item) as f:
#         data = json.load(f)
#         # if data["rounds"][0]["matchType"]["name"]!="CustomGameOnline" or data["rounds"][0]["matchType"]["name"]!="CustomGameLocal":
#         #     notcomplist.append(item)
#         #     jsonlist.remove(item)

for item in notcomplist:
    os.rename(path+'\\'+item, path+'\\Other'+item) #move dissected files to a new folder

for item in jsonlist: #ask if they want to read the file
    while True:
        print(jsonlist)
        ask = input('Do you want to read '+item+'? (Y or N) ').lower()
        checkask = ['y','yes','n','no']
        if ask not in checkask:
            print('Invalid input')
            continue
        elif ask == 'y' or ask == 'yes':
            numberofmatches+=1
            matcheslist.append(item)
            print(f'{item} loaded as Match{numberofmatches}')
            break
        elif ask == 'n' or ask == 'no':
            print('Not loaded')
            break

names=[]
for matchiteration in range(int(numberofmatches)):
    path2 = path+'\\'+matcheslist[matchiteration]
    with open(path2) as f:
        data = json.load(f)
    stats=pd.DataFrame() #create a list to hold all the player stats
    rounds = data["rounds"]
    mapname= rounds[0]["map"]["name"]
    team1 = rounds[0]["teams"][0]["name"]
    team2 = rounds[0]["teams"][1]["name"]
    team1players=[]
    team2players=[]
    mapname=rounds[0]["map"]["name"]
    globals()[f'{team1}_Match_{matchiteration}_Score']=rounds[len(rounds)-1]["teams"][0]["score"]
    globals()[f'{team2}_Match_{matchiteration}_Score']=rounds[len(rounds)-1]["teams"][1]["score"]
    for i in range(10):
        if rounds[0]["players"][i]["teamIndex"]==0:
            team1players.append(rounds[0]["players"][i]["username"])
        elif rounds[0]["players"][i]["teamIndex"]==1:
            team2players.append(rounds[0]["players"][i]["username"])
    for playernum in range(10):
        name=rounds[0]["players"][playernum]["username"]
        if name not in names:
            names.append(name) #add player name to the list
        ops=tuple() 
        entrykills=0
        entrydeaths=0 
        kost=0
        roundssrv=0
        objplays=int(0)
        vx=0
        kills_rounds=[]
        hsp_rounds=[]
        for roundnum in range(len(rounds)):
            planted=False
            defused=False
            traded=False
            kill=False
            survive=False
            roundwl="L"
            for i in range(10): #find the player number in the round and get operator
                if rounds[roundnum]["players"][i]["username"]==name:
                    playernumber_round=i
                    op=rounds[roundnum]["players"][i]["operator"]["name"] #get the operator for the round
                    teamnumber=rounds[roundnum]["players"][playernumber_round]["teamIndex"]
                    # if rounds[roundnum]["players"][i]["teamIndex"]==0 and rounds[roundnum]["teams"][0]["won"]== True: #if the player is on team 1 and they won
                    #     roundwl="W"
                    # elif rounds[roundnum]["players"][i]["teamIndex"]==1 and rounds[roundnum]["teams"][1]["won"]== True: #if the player is on team 2 and they won
                    #     roundwl="W"
                    if roundnum == 0:
                        if rounds[roundnum]["teams"][teamnumber]["score"]==1:
                            roundwl="W"
                    elif rounds[roundnum]["teams"][teamnumber]["score"]==rounds[roundnum-1]["teams"][teamnumber]["score"]+1: #if the score is 1 higher than the previous round
                        roundwl="W"
                    ops+=tuple([[op, roundwl]]) #add the operator to the list
                    break
            for i in range(len(rounds[roundnum]["matchFeedback"])): #calculate entry kills and deaths
                if rounds[roundnum]["matchFeedback"][i]["type"]["name"]=="Kill":
                    if rounds[roundnum]["matchFeedback"][i]["username"]==name:
                        entrykills+=1 #will continue to count up between games
                        break
                    elif rounds[roundnum]["matchFeedback"][i]["target"]==name:
                        entrydeaths+=1 #will continue to count up between games
                        break
                    else:
                        break
            for i2 in range(len(rounds[roundnum]["matchFeedback"])): #calculate plants 
                if rounds[roundnum]["matchFeedback"][i2]["type"]["name"]=="DefuserPlantComplete" and rounds[roundnum]["matchFeedback"][i2]["username"]==name:
                    planted=True
                    objplays+=1 #will continue to count up between games
                    break
            for defcheck in range(len(rounds[roundnum]["matchFeedback"])): #calculate defuses 
                if rounds[roundnum]["matchFeedback"][defcheck]["type"]["name"]=="DefuserDisableComplete" and rounds[roundnum]["matchFeedback"][defcheck]["username"]==name:
                    defused=True 
                    objplays+=1 #will continue to count up between games           
                    break
            if rounds[roundnum]["stats"][playernumber_round]["kills"]>0: #if gets a kill, add to kost
                kill=True
            for i4 in range(len(rounds[roundnum]["matchFeedback"])): #calculate trades for kost need check
                if rounds[roundnum]["matchFeedback"][i4]["type"]["name"]=="Kill" and rounds[roundnum]["matchFeedback"][i4]["target"]==name:
                    tod=rounds[roundnum]["matchFeedback"][i4]["timeInSeconds"]
                    killer=rounds[roundnum]["matchFeedback"][i4]["username"]     
                    for tradecheck in range(i4+1,len(rounds[roundnum]["matchFeedback"])): #within 8 seconds
                        if rounds[roundnum]["matchFeedback"][tradecheck]["type"]["name"]=="Kill" and rounds[roundnum]["matchFeedback"][tradecheck]["target"]==killer and tod-rounds[roundnum]["matchFeedback"][tradecheck]["timeInSeconds"]<=3:
                            traded=True
                            break
            if rounds[roundnum]["stats"][playernumber_round]["died"]==False:
                survive=True
                roundssrv+=1 #will continue to count up between games
            if kill==True or defused==True or planted==True or traded==True or survive==True:
                kost+=1 #will continue to count up between games
            if rounds[roundnum]["stats"][playernumber_round]["died"]==False: #1vX, if this player survived
                if roundnum == 0:
                    if rounds[roundnum]["teams"][teamnumber]["score"]==1:
                        if rounds[roundnum]["teams"][teamnumber]["role"]=="Defense": #defendin, players 1-5
                            for possibleteammates in range(5):
                                if rounds[roundnum]["stats"][possibleteammates]["died"]==False and possibleteammates!=playernumber_round: #if there is a surviving teammate that is not the player
                                    #do once invalid
                                    possibleteammates=0
                                    break
                                elif possibleteammates==4: #gone through every teammate and they all died
                                    vx+=1 #will continue to count up between games
                                else: #if still valid
                                    continue
                        else: #attacking, players 6-10
                            for possibleteammates in range(6,10):
                                if rounds[roundnum]["stats"][possibleteammates]["died"]==False and possibleteammates!=playernumber_round:
                                    #do once invalid
                                    possibleteammates=0
                                    break
                                elif possibleteammates==9: #gone through every teammate and they all died
                                    vx+=1 #will continue to count up between games
                                else: #if still valid
                                    continue
                if rounds[roundnum]["teams"][teamnumber]["score"]==rounds[roundnum-1]["teams"][teamnumber]["score"]+1: #if the score is 1 higher than the previous round
                    if rounds[roundnum]["teams"][teamnumber]["role"]=="Defense": #defendin, players 1-5
                        for possibleteammates in range(5): #work through all players on the team
                            if rounds[roundnum]["stats"][possibleteammates]["died"]==False and possibleteammates!=playernumber_round:
                                #do once invalid
                                possibleteammates=0
                                break
                            elif possibleteammates==4: #gone through every teammate and they all died
                                vx+=1 #will continue to count up between games
                            else: #if still valid
                                continue
                    else: #attacking, players 6-10
                        for possibleteammates in range(6,10):
                            if rounds[roundnum]["stats"][possibleteammates]["died"]==False and possibleteammates!=playernumber_round:
                                #do once invalid
                                possibleteammates=0
                                break
                            elif possibleteammates==9: #gone through every teammate and they all died
                                vx+=1 #will continue to count up between games
                            else: #if still valid
                                continue
                elif "1vX" in rounds[roundnum]["stats"][playernumber_round]:
                    vx+=1 
            kills_rounds.append(rounds[roundnum]["stats"][playernumber_round]["kills"]) 
            hsp_rounds.append(rounds[roundnum]["stats"][playernumber_round]["headshotPercentage"])
        if matchiteration==0 or f'{name}_Stats' not in globals():
            kills=int(data["stats"][playernum]["kills"])
            deaths=int(data["stats"][playernum]["deaths"])
            roundsplayed=int(data["stats"][playernum]["rounds"])
            hsp=data["stats"][playernum]["headshotPercentage"]
            globals()[f'{name}_Stats']=pd.Series(kills, index=["Kills"]) #create a dataframe for each player, starts with kills so it doesnt concat a blank
            globals()[f'{name}_Stats']=pd.concat([globals()[f'{name}_Stats'], pd.Series(deaths, index=["Deaths"])]) #add deaths to the stat var
            globals()[f'{name}_Stats']=pd.concat([globals()[f'{name}_Stats'], pd.Series(roundsplayed, index=["Rounds"])]) #add rounds to the stat var
            globals()[f'{name}_Stats']=pd.concat([globals()[f'{name}_Stats'], pd.Series(int(objplays), index=["OBJ"])]) #add obj to the stat var
            globals()[f'{name}_StatsHS']=pd.Series(hsp, index=["HS%"]) #add HS% to seperate var
            globals()[f'{name}_Stats']=pd.concat([globals()[f'{name}_Stats'], pd.Series(int(kost), index=["KOST"])]) #add KOST rounds to the stat var
            globals()[f'{name}_Stats']=pd.concat([globals()[f'{name}_Stats'], pd.Series(int(roundssrv), index=["SRV"])]) #add SRV to the stat var
            globals()[f'{name}_Stats']=pd.concat([globals()[f'{name}_Stats'], pd.Series(int(vx), index=["1vx"])]) #add 1vx to the stat var
            globals()[f'{name}_Stats']=pd.concat([globals()[f'{name}_Stats'], pd.Series(int(entrykills), index=["Entry Kills"])]) #add entry kills to the stat var
            globals()[f'{name}_Stats']=pd.concat([globals()[f'{name}_Stats'], pd.Series(int(entrydeaths), index=["Entry Deaths"])]) #add entry deaths to the stat var
            globals()[f'{name}_Ops']=ops #create a global tuple for each player's ops
            globals()[f'{name}_KillList']=kills_rounds #create a global list for each player's kills per round
            globals()[f'{name}_HSList']=hsp_rounds #create a global list for each player's hs% per round
            if name in team1players:
                globals()[f'{name}_Stats']=pd.concat([globals()[f'{name}_Stats'], pd.Series(team1, index=["Team"])])
            elif name in team2players:
                globals()[f'{name}_Stats']=pd.concat([globals()[f'{name}_Stats'], pd.Series(team2, index=["Team"])]) 
        else:
            globals()[f'{name}_Stats']["Kills"]+=data["stats"][playernum]["kills"]
            globals()[f'{name}_Stats']["Deaths"]+=data["stats"][playernum]["deaths"]
            globals()[f'{name}_Stats']["Rounds"]+=data["stats"][playernum]["rounds"]
            globals()[f'{name}_Stats']["OBJ"]+=int(objplays)
            globals()[f'{name}_StatsHS']["HS%"]=float(((globals()[f'{name}_StatsHS']["HS%"]*(matchiteration))+data["stats"][playernum]["headshotPercentage"]) / (matchiteration+1))
            globals()[f'{name}_Stats']["KOST"]+=int(kost)
            globals()[f'{name}_Stats']["SRV"]+=int(roundssrv)
            globals()[f'{name}_Stats']["1vx"]+=int(vx)
            globals()[f'{name}_Stats']["Entry Kills"]+=int(entrykills)
            globals()[f'{name}_Stats']["Entry Deaths"]+=int(entrydeaths)
            globals()[f'{name}_Ops'] += ops #add ops to the list
            globals()[f'{name}_KillList'] += kills_rounds #add kills per round to the list
            globals()[f'{name}_HSList'] += hsp_rounds #add hs% per round to the list
    #os.rename(path2, "Jsons\\Dissected\\"+matcheslist[matchiteration]) #move dissected files to a new folder
    #os.rename(path2, "Jsons\\Dissected\\"+f'{team1}_{team2}_Match_{matchiteration}.json') #rename the file to include the teams
    if dissect==True: #if a new file was dissected, rename the files accordingly
        if f'{team1}_{team2}_Match_{matchiteration+1}.json' not in os.listdir("Jsons"): #if the file does not exist
            os.rename(path2, "Jsons\\"+f'{team1}_{team2}_Match_{matchiteration+1}.json') #rename the file to include the teams
        else: #add a number to the end of the filename, incrementing until it is unique
            i=2
            while True:
                if f'{team1}_{team2}_Match_{matchiteration+1}_{i}.json' not in os.listdir("Jsons"):
                    os.rename(path2, "Jsons\\"+f'{team1}_{team2}_Match_{matchiteration+1}_{i}.json')
                    break
                else:
                    i+=1
for name in names: #create actual stats
    kills=globals()[f'{name}_Stats']["Kills"]
    deaths=globals()[f'{name}_Stats']["Deaths"]
    if deaths==0:
        deaths=1
    roundsplayed=globals()[f'{name}_Stats']["Rounds"]
    objplays=globals()[f'{name}_Stats']["OBJ"]
    hsp=globals()[f'{name}_StatsHS']["HS%"]
    kost=globals()[f'{name}_Stats']["KOST"]
    roundssrv=globals()[f'{name}_Stats']["SRV"]
    vx=globals()[f'{name}_Stats']["1vx"]
    entrykills=globals()[f'{name}_Stats']["Entry Kills"]
    entrydeaths=globals()[f'{name}_Stats']["Entry Deaths"]
    playerteam=globals()[f'{name}_Stats']["Team"]
    kd=pd.Series(round(kills/deaths,2), index=["K/D"])
    kd2=pd.Series(f'{kills}-{deaths} ({kills-deaths})', index=["K/D (+/-)"])
    kpr=pd.Series(round(kills/roundsplayed,2), index=["KPR"])
    rating=round(((0.8*kills)/roundsplayed)+((0.3*kost)/roundsplayed)+((0.5*roundssrv)/roundsplayed)+((0.8*objplays)/roundsplayed)+((1.0*vx)/roundsplayed)+(0.2*(entrykills-entrydeaths)/roundsplayed),2)
    entrystat=pd.Series(f'{entrykills} - {entrydeaths} ({entrykills-entrydeaths})', index=["Entry (+/-)"])
    koststat=pd.Series(round(kost/roundsplayed, 2), index=["KOST"])
    srvstat=pd.Series(round(roundssrv/roundsplayed, 2), index=["SRV"])
    vxstat=pd.Series(round(vx, 2), index=["1vx"])
    obj=pd.Series(objplays, index=["OBJ"])
    combostat=pd.Series(rating, index=["Rating"]) #create a dataframe for each player starting with rating
    combostat=pd.concat([combostat, kd]) #add K/D to the list
    combostat=pd.concat([combostat, kd2]) #add kd+/- to the list
    combostat=pd.concat([combostat, entrystat]) #add entry kills to the list
    combostat=pd.concat([combostat, koststat]) #add KOST to the list
    combostat=pd.concat([combostat, kpr]) #add kpr to the list
    combostat=pd.concat([combostat, srvstat]) #add SRV to the list
    combostat=pd.concat([combostat, vxstat]) #add 1vx to the list
    combostat=pd.concat([combostat, obj]) #add obj to the list
    combostat=pd.concat([combostat, pd.Series(round(hsp,2), index=["HS%"])]) #add HS% to the list
    combostat=pd.concat([combostat, pd.Series(roundsplayed, index=["Rounds"])]) #add rounds to the list
    combostat=pd.concat([combostat, pd.Series(playerteam, index=["Team"])]) #add team to the list
    if stats.empty: #for subs, puts players in their teams 
        stats=pd.concat([stats, pd.DataFrame(combostat, columns=[name])], axis=1)
    elif stats.iloc[11,-1]==combostat["Team"]: #if the last team is the same as the current team
        stats=pd.concat([stats, pd.DataFrame(combostat, columns=[name])], axis=1)
    elif combostat["Team"] not in stats.iloc[11,:].values: #if the team is not in the list
        stats=pd.concat([stats, pd.DataFrame(combostat, columns=[name])], axis=1)
    else:
        for i in range(len(stats)):
            if stats.iloc[11,i]!=combostat["Team"]: #find location of first instance of the team
                stats=pd.concat([stats.iloc[:,:i], pd.DataFrame(combostat, columns=[name]), stats.iloc[:,i:]], axis=1)
                break

try:
    stats=stats.T
except:
    print("Error: No jsons found, either put one in the Jsons folder or dissect a new match. Exiting...")
    exit()

print(stats)

if f'{team1}_{team2}.xlsx' not in os.listdir("Outputs"): #if the file does not exist
    stats.to_excel("Outputs\\"+f'{team1}_{team2}.xlsx') #export to excel
    filename=f'{team1}_{team2}.xlsx'
else: #add a number to the end of the filename, incrementing until it is unique
    i=2
    while True:
        if f'{team1}_{team2}_{i}.xlsx' not in os.listdir("Outputs"):
            stats.to_excel("Outputs\\"+f'{team1}_{team2}_{i}.xlsx')
            filename=f'{team1}_{team2}_{i}.xlsx'
            break
        else:
            i+=1

#doesnt move jsons, uncomment

wb = openpyxl.load_workbook(f'Outputs\\{filename}') #add finals scores to the excel
ws = wb.active
ws.delete_cols(13)
ws.insert_rows(1,4)
for i in range(numberofmatches):
    ws.cell(row=1, column=1+(i*3)).value = f'Match {i+1}'
    ws.cell(row=1, column=1+(i*3)).font = Font(bold=True)
    ws.cell(row=1, column=1+(i*3)).border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    ws.cell(row=1, column=2+(i*3)).value = "Score"
    ws.cell(row=1, column=2+(i*3)).font = Font(bold=True)
    ws.cell(row=1, column=2+(i*3)).border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    ws.cell(row=2, column=1+(i*3)).value = f'{team1}'
    ws.cell(row=2, column=1+(i*3)).font = Font(bold=True)
    ws.cell(row=2, column=1+(i*3)).border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    ws.cell(row=2, column=2+(i*3)).value = int(f'{globals()[f"{team1}_Match_{i}_Score"]}')
    ws.cell(row=2, column=2+(i*3)).border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    ws.cell(row=3, column=1+(i*3)).value = f'{team2}'
    ws.cell(row=3, column=1+(i*3)).font = Font(bold=True)
    ws.cell(row=3, column=1+(i*3)).border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    ws.cell(row=3, column=2+(i*3)).value = int(f'{globals()[f"{team2}_Match_{i}_Score"]}')
    ws.cell(row=3, column=2+(i*3)).border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
wb.save(f'Outputs\\{filename}')

#make matplotlib plots not display
plt.ioff()

wb = openpyxl.load_workbook(f'Outputs\\{filename}')
for name in names:
    sheets=wb.sheetnames
    if name not in sheets:
        wb.create_sheet(name)
    ws = wb[name]
    for i in range(len(globals()[f'{name}_Ops'])):
        ws.cell(row=i+1, column=1).value = globals()[f'{name}_Ops'][i][0]
        ws.cell(row=i+1, column=2).value = globals()[f'{name}_Ops'][i][1]
    
    plt.bar(range(len(globals()[f'{name}_KillList'])),globals()[f'{name}_KillList'])
    plt.title(f'{name} Kills per round')
    plt.xlabel("Rounds")
    plt.ylabel("Kills")
    #add black border around the plot
    plt.gca().spines['top'].set_color('black')
    plt.gca().spines['right'].set_color('black')
    plt.gca().spines['bottom'].set_color('black')
    plt.gca().spines['left'].set_color('black')

    plt.savefig(f'Outputs\\{name}_KPR.png')
    img = openpyxl.drawing.image.Image(f'Outputs\\{name}_KPR.png')
    ws.add_image(img, 'D1')
    plt.close()

    plt.plot(globals()[f'{name}_HSList'])  
    plt.title(f'{name} HS% per round')
    plt.xlabel("Rounds")
    plt.ylabel("HS%")

    plt.savefig(f'Outputs\\{name}_HS.png')
    img = openpyxl.drawing.image.Image(f'Outputs\\{name}_HS.png')
    ws.add_image(img, 'O1')
    plt.close()

    wb.save(f'Outputs\\{filename}')
for file in os.listdir("Outputs"):
    if file.endswith(".png"):
        os.remove(f"Outputs\\{file}")